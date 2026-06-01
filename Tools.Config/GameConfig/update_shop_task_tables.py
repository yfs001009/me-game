from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE = Path(__file__).resolve().parent / "Datas"


def save_workbook(name, rows):
    path = BASE / name
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


def upsert_table(ws, full_name, value_type, input_file, index, comment):
    for row in range(4, ws.max_row + 1):
        if ws.cell(row, 2).value == full_name:
            ws.cell(row, 3).value = value_type
            ws.cell(row, 4).value = True
            ws.cell(row, 5).value = input_file
            ws.cell(row, 6).value = index
            ws.cell(row, 9).value = comment
            return
    ws.append([None, full_name, value_type, True, input_file, index, None, None, comment, None])


def remove_table(ws, full_name):
    for row in range(ws.max_row, 3, -1):
        if ws.cell(row, 2).value == full_name:
            ws.delete_rows(row, 1)


def update_tables():
    path = BASE / "__tables__.xlsx"
    wb = load_workbook(path)
    ws = wb.active

    remove_table(ws, "battle.TbShop")
    remove_table(ws, "battle.TbShopGoods")
    remove_table(ws, "outgame.TbOutgameShop")
    remove_table(ws, "outgame.TbOutgameShopGoods")
    remove_table(ws, "outgame.TbOutgameTask")

    upsert_table(ws, "open.TbOpenFeature", "OpenFeatureConfig", "open_feature.xlsx", "featureId", "统一开放定义表：活动、功能、限时入口、任务和商店共用")
    upsert_table(ws, "shop.TbShop", "ShopConfig", "shop.xlsx", "shopId", "局外商店表：常驻、活动和限时商店")
    upsert_table(ws, "shop.TbShopGoods", "ShopGoodsConfig", "shop_goods.xlsx", "goodsId", "局外商店商品表")
    upsert_table(ws, "task.TbTask", "TaskConfig", "task.xlsx", "taskId", "局外任务表：每日、周常、成就和活动任务")
    upsert_table(ws, "battle.TbBattleShop", "BattleShopConfig", "battle_shop.xlsx", "shopId", "局内商店表：Tiled shop 层对象通过 shop_id 引用")
    upsert_table(ws, "battle.TbBattleShopGoods", "BattleShopGoodsConfig", "battle_shop_goods.xlsx", "goodsId", "局内商店商品表：定义本局售卖物品、价格和解锁条件")

    wb.save(path)


save_workbook(
    "open_feature.xlsx",
    [
        ["##var", "featureId", "name", "category", "openConditionType", "openParam", "startTime", "endTime", "isEnabled", "comment"],
        ["##type", "string", "string", "string", "string", "string", "string", "string", "bool", "string"],
        ["##group", None, None, None, None, None, None, None, None, None],
        ["##", "开放ID；其他表通过该字段关联", "显示名", "分类：Daily/Shop/Activity/System", "Always/TimeRange/Level/ServerSwitch", "开放参数，按类型解释", "UTC开始时间，空表示不限", "UTC结束时间，空表示不限", "总开关", "说明"],
        [None, "DailyActive", "每日活跃", "Daily", "Always", "", "", "", True, "每日任务和每日活跃入口共用"],
        [None, "Shop.Normal", "常驻局外商店", "Shop", "Always", "", "", "", True, "常驻局外商店"],
        [None, "Activity.Launch", "开服活动", "Activity", "Always", "", "", "", True, "开服活动任务和活动商店共用"],
    ],
)

save_workbook(
    "shop.xlsx",
    [
        ["##var", "shopId", "shopName", "shopType", "featureId", "activityId", "goodsGroupId", "refreshGroup", "comment"],
        ["##type", "int", "string", "string", "string", "string", "int", "string", "string"],
        ["##group", None, None, None, None, None, None, None, None],
        ["##", "商店ID", "商店显示名", "商店类型：Normal/Activity/Limited", "关联 open_feature.featureId", "活动ID，仅用于归类筛选", "商品组ID；关联 shop_goods.goodsGroupId", "刷新组：Daily/Weekly/Activity.xxx/Permanent", "说明"],
        [None, 2001, "常驻补给商店", "Normal", "Shop.Normal", "", 2001, "Permanent", "局外常驻商品"],
        [None, 3001, "开服活动商店", "Activity", "Activity.Launch", "Launch", 3001, "Activity.Launch", "开服活动期间开放"],
    ],
)

save_workbook(
    "shop_goods.xlsx",
    [
        ["##var", "goodsId", "goodsGroupId", "itemId", "itemName", "price", "currency", "buyLimit", "rewardItemCount", "unlockRule", "comment", "effectDesc"],
        ["##type", "int", "int", "int", "string", "int", "string", "int", "int", "string", "string", "string"],
        ["##group", None, None, None, None, None, None, None, None, None, None, None],
        ["##", "商品ID", "商品组ID；关联 shop.goodsGroupId", "奖励道具ID", "商品显示名", "价格", "货币类型：Gold/Diamond/EventToken", "购买上限，0表示不限", "每次购买获得数量", "解锁条件，空表示默认解锁", "说明", "效果说明"],
        [None, 200101, 2001, 1001, "普通抽奖券", 200, "Gold", 10, 1, "", "常驻补给", "获得 1 张普通抽奖券"],
        [None, 300101, 3001, 1301, "活动兑换碎片", 10, "EventToken", 99, 10, "", "开服活动商品", "获得 10 个活动兑换碎片"],
    ],
)

save_workbook(
    "task.xlsx",
    [
        ["##var", "taskId", "taskType", "featureId", "activityId", "title", "description", "progressKey", "target", "refreshGroup", "rewardCurrencyId", "rewardCurrencyAmount", "rewardItemId", "rewardItemCount", "comment"],
        ["##type", "int", "string", "string", "string", "string", "string", "string", "long", "string", "int", "long", "int", "int", "string"],
        ["##group", None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ["##", "任务ID", "任务类型：Daily/Weekly/Achievement/Activity", "关联 open_feature.featureId", "活动ID，仅用于归类筛选", "标题", "描述", "进度Key，由服务端行为打点", "目标值", "刷新组", "奖励货币ID，0表示无", "奖励货币数量", "奖励道具ID，0表示无", "奖励道具数量", "说明"],
        [None, 1001, "Daily", "DailyActive", "", "每日登录", "登录游戏 1 次。", "Login.Count", 1, "Daily", 1, 100, 0, 0, "每日活跃"],
        [None, 1002, "Daily", "DailyActive", "", "每日抽奖", "完成 1 次抽奖。", "Lottery.Any.DrawCount", 1, "Daily", 0, 0, 1001, 1, "每日活跃"],
        [None, 2001, "Activity", "Activity.Launch", "Launch", "开服采购", "在任意局外商店购买 1 次商品。", "Shop.Buy.Count", 1, "Activity.Launch", 3, 10, 0, 0, "开服活动"],
    ],
)

save_workbook(
    "battle_shop.xlsx",
    [
        ["##var", "shopId", "shopName", "shopType", "goodsGroupId", "ownerCamp", "comment", "effectDesc"],
        ["##type", "int", "string", "string", "int", "string", "string", "string"],
        ["##group", None, None, None, None, None, None, None],
        ["##", "商店ID", "商店显示名", "商店类型，例如 TrollWeapon", "商品组ID；关联 battle_shop_goods.goodsGroupId", "可使用阵营，例如 Troll", "中文说明", "效果说明"],
        [None, 1001, "巨魔武器商店", "TrollWeapon", 1001, "Troll", "Tiled shop 层对象填写 shop_id=1001 时生成。", "巨魔在局内购买武器和基础装备。"],
    ],
)

save_workbook(
    "battle_shop_goods.xlsx",
    [
        ["##var", "goodsId", "goodsGroupId", "itemId", "itemName", "price", "currency", "unlockRule", "comment", "effectDesc"],
        ["##type", "int", "int", "int", "string", "int", "string", "string", "string", "string"],
        ["##group", None, None, None, None, None, None, None, None, None],
        ["##", "商品ID", "商品组ID；关联 battle_shop.goodsGroupId", "局内物品ID；后续关联 weapon/equipment 表", "商品显示名", "价格", "局内货币类型，例如 Gold/Wood", "解锁条件，空表示默认解锁", "中文说明", "效果说明"],
        [None, 100101, 1001, 2001, "粗制战斧", 100, "Gold", "", "巨魔武器商店默认商品。", "购买后提高巨魔近战攻击能力。"],
        [None, 100102, 1001, 2002, "骨质护甲", 120, "Gold", "", "巨魔武器商店默认商品。", "购买后提高巨魔生存能力。"],
    ],
)

for stale in ["outgame_shop.xlsx", "outgame_shop_goods.xlsx", "outgame_task.xlsx"]:
    path = BASE / stale
    if path.exists():
        path.unlink()

update_tables()
