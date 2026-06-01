from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE = Path(__file__).resolve().parent / "Datas"


def save_workbook(name, rows):
    path = BASE / name
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


def update_tables():
    path = BASE / "__tables__.xlsx"
    wb = load_workbook(path)
    ws = wb.active

    rows_by_full_name = {
        ws.cell(row, 2).value: row
        for row in range(4, ws.max_row + 1)
        if ws.cell(row, 2).value
    }

    rows = {
        "asset.TbCurrency": [
            None,
            "asset.TbCurrency",
            "CurrencyConfig",
            True,
            "currency.xlsx",
            "currencyId",
            None,
            None,
            "局外货币配置表：金币、钻石、活动币等纯数值资产",
            None,
            None,
        ],
        "asset.TbItem": [
            None,
            "asset.TbItem",
            "ItemConfig",
            True,
            "item.xlsx",
            "itemId",
            None,
            None,
            "局外道具配置表：票券、体验卡、碎片、增益卡、宝箱和材料",
            None,
            None,
        ],
    }

    for full_name, values in rows.items():
        row_index = rows_by_full_name.get(full_name)
        if row_index is None:
            row_index = ws.max_row + 1
        for col, value in enumerate(values, 1):
            ws.cell(row_index, col).value = value

    wb.save(path)


save_workbook(
    "currency.xlsx",
    [
        ["##var", "currencyId", "code", "name", "iconAsset", "isPremium", "sortOrder", "description"],
        ["##type", "int", "string", "string", "string", "bool", "int", "string"],
        ["##group", None, None, None, None, None, None, None],
        ["##", "货币ID", "货币代码", "显示名", "图标资源名", "是否付费货币", "排序权重", "中文说明"],
        [None, 1, "Gold", "金币", "Icon_Currency_Gold", False, 10, "局外通用软货币，用于购买和升级。"],
        [None, 2, "Diamond", "钻石", "Icon_Currency_Diamond", True, 20, "付费或稀有货币，用于高级购买。"],
        [None, 3, "EventToken", "活动币", "Icon_Currency_EventToken", False, 30, "活动兑换用限时货币。"],
    ],
)

save_workbook(
    "item.xlsx",
    [
        [
            "##var",
            "itemId",
            "itemType",
            "name",
            "quality",
            "maxStack",
            "useType",
            "effectType",
            "effectParam",
            "iconAsset",
            "description",
        ],
        [
            "##type",
            "int",
            "string",
            "string",
            "int",
            "int",
            "string",
            "string",
            "string",
            "string",
            "string",
        ],
        ["##group", None, None, None, None, None, None, None, None, None, None],
        [
            "##",
            "道具ID",
            "道具类型：Ticket/Consumable/Fragment/Chest/Trial/Material",
            "显示名",
            "品质",
            "堆叠上限",
            "使用方式：None/UseDirectly/OpenChest/ActivateBuff/Exchange",
            "效果类型",
            "效果参数",
            "图标资源名",
            "中文说明",
        ],
        [None, 1001, "Ticket", "普通抽奖券", 2, 999, "UseDirectly", "Lottery", "Normal", "Icon_Item_LotteryTicket", "用于普通奖池抽奖一次。"],
        [None, 1002, "Ticket", "高级抽奖券", 3, 999, "UseDirectly", "Lottery", "Premium", "Icon_Item_PremiumLotteryTicket", "用于高级奖池抽奖一次。"],
        [None, 1101, "Consumable", "双倍经验卡", 2, 99, "ActivateBuff", "DoubleExp", "3600", "Icon_Item_DoubleExp", "使用后获得 1 小时双倍经验。"],
        [None, 1102, "Consumable", "双倍金币卡", 2, 99, "ActivateBuff", "DoubleGold", "3600", "Icon_Item_DoubleGold", "使用后获得 1 小时双倍金币。"],
        [None, 1201, "Trial", "精灵游侠体验卡", 3, 99, "UseDirectly", "CharacterTrial", "1002:86400", "Icon_Item_ElfRangerTrial", "使用后获得精灵游侠 1 天体验。"],
        [None, 1202, "Trial", "雾行者体验卡", 3, 99, "UseDirectly", "CharacterTrial", "2002:86400", "Icon_Item_TrollMistTrial", "使用后获得雾行者 1 天体验。"],
        [None, 1301, "Fragment", "精灵游侠碎片", 2, 9999, "Exchange", "CharacterShard", "1002", "Icon_Item_ElfRangerShard", "集齐后可兑换精灵游侠。"],
        [None, 1302, "Fragment", "防御塔卡片碎片", 2, 9999, "Exchange", "BuildingCardShard", "1003", "Icon_Item_TowerCardShard", "集齐后可兑换防御塔建筑卡。"],
        [None, 1401, "Chest", "新手补给箱", 2, 99, "OpenChest", "RewardPackage", "NewPlayerSupply", "Icon_Item_NewPlayerChest", "打开后获得新手阶段常用奖励。"],
        [None, 1501, "Material", "通用升级石", 2, 9999, "None", "UpgradeMaterial", "Universal", "Icon_Item_UpgradeStone", "用于后续角色或建筑卡升级。"],
    ],
)

update_tables()
