from pathlib import Path

from openpyxl import load_workbook


base = Path(__file__).resolve().parent / "Datas"
for name in ["__tables__.xlsx", "shop.xlsx", "shop_goods.xlsx", "__beans__.xlsx", "__enums__.xlsx"]:
    path = base / name
    wb = load_workbook(path, data_only=False)
    print()
    print(name, wb.sheetnames)
    ws = wb.active
    for row in range(1, min(ws.max_row, 8) + 1):
        values = [ws.cell(row, col).value for col in range(1, min(ws.max_column, 10) + 1)]
        print(row, values)
