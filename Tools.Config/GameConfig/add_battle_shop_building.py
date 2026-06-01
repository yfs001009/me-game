from openpyxl import load_workbook


path = "Datas/building.xlsx"
building_id = 401
row_values = [
    None,
    building_id,
    "局内商店",
    "Shop",
    "Building_BattleShop",
    1,
    1,
    1,
    1,
    0,
    0,
    False,
    False,
    False,
    "地图商店建筑",
    "巨魔选中后可打开局内商店",
]

workbook = load_workbook(path)
worksheet = workbook.active

id_column = 2
target_row = None
for row in range(5, worksheet.max_row + 1):
    if worksheet.cell(row=row, column=id_column).value == building_id:
        target_row = row
        break

if target_row is None:
    target_row = worksheet.max_row + 1

for column, value in enumerate(row_values, start=1):
    worksheet.cell(row=target_row, column=column).value = value

workbook.save(path)
print(f"upserted battle shop building: row={target_row}, buildingId={building_id}")
