from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent / "Datas"


def save(workbook_name, updates):
    path = BASE / workbook_name
    wb = load_workbook(path)
    ws = wb.active
    for cell, value in updates.items():
        ws[cell] = value
    wb.save(path)


save(
    "character.xlsx",
    {
        "B4": "角色ID",
        "C4": "角色分类：Hero/Ghost",
        "D4": "种族/阵营：Elf/Troll",
        "E4": "角色名",
        "F4": "技能ID",
        "G4": "技能名称",
        "H4": "技能描述",
        "I4": "图标资源名",
        "J4": "预制体资源名",
        "K4": "是否初始解锁",
        "L4": "排序权重",
        "M4": "角色说明",
        "E5": "默认精灵",
        "G5": "轻盈步伐",
        "H5": "移动速度提高 5%。",
        "M5": "首版默认精灵角色，适合新手防守和建造。",
        "E6": "精灵游侠",
        "G6": "远行者",
        "H6": "建造范围和维修范围小幅提高。",
        "M6": "偏向机动和支援的精灵角色，后续通过商店解锁。",
        "E7": "破墙者",
        "G7": "粉碎",
        "H7": "对建筑伤害提高 30%。",
        "M7": "首版默认巨魔角色，擅长正面拆墙。",
        "E8": "雾行者",
        "G8": "雾隐",
        "H8": "短暂隐身 1 秒。",
        "M8": "偏向绕后偷袭的巨魔角色，后续通过商店解锁。",
    },
)

save(
    "__tables__.xlsx",
    {
        "I12": "角色配置表：精灵/巨魔角色、技能、资源和解锁状态",
    },
)
